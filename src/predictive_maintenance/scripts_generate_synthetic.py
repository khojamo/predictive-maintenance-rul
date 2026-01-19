from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
import sys

import numpy as np
import pandas as pd

from predictive_maintenance.feature_catalog import (
    CANONICAL_FEATURES,
    ID_COL,
    TIME_COL,
    LABEL_FAIL_WITHIN_H,
    LABEL_FAILURE_CYCLE,
    LABEL_RUL,
)

EXCEL_MAX_ROWS = 1_048_576  # hard Excel limit per sheet


@dataclass(frozen=True)
class GeneratorConfig:
    n_units: int
    cycles: int
    horizon: int
    chunk_units: int
    seed: int


def _feature_spec(name: str) -> tuple[float, float, float, float, float | None]:
    """Return (mean, std, trend, noise, clip_low)."""
    if name == "operating_mode":
        return 0.0, 0.0, 0.0, 0.0, None
    if name == "runtime_hours":
        return 0.0, 0.0, 0.0, 0.0, 0.0

    if "temperature" in name:
        return 60.0, 5.0, 18.0, 1.0, 0.0
    if "pressure" in name or "pressure_ratio" in name:
        return 30.0, 3.0, 8.0, 0.5, 0.0
    if "vibration" in name:
        return 0.5, 0.1, 1.0, 0.1, 0.0
    if "speed" in name:
        return 3000.0, 80.0, -200.0, 20.0, 0.0
    if "torque" in name:
        return 150.0, 20.0, 30.0, 5.0, 0.0
    if "power_output" in name:
        return 500.0, 50.0, -80.0, 10.0, 0.0
    if "current" in name:
        return 50.0, 5.0, 8.0, 2.0, 0.0
    if "voltage" in name:
        return 400.0, 10.0, 0.0, 2.0, 0.0
    if "efficiency" in name:
        return 0.9, 0.02, -0.15, 0.01, 0.0
    if "power_factor" in name:
        return 0.95, 0.02, -0.1, 0.01, 0.0
    if "load_percent" in name:
        return 70.0, 10.0, 0.0, 5.0, 0.0
    if "flow_rate" in name or name.endswith("flow_rate") or "flow" in name:
        return 100.0, 10.0, -10.0, 5.0, 0.0
    if "valve_position" in name or "actuator_position" in name:
        return 50.0, 15.0, 5.0, 5.0, 0.0
    if "setpoint" in name:
        return 100.0, 5.0, 0.0, 2.0, 0.0
    if "error_signal" in name:
        return 0.0, 1.0, 3.0, 1.0, None
    if "energy_consumption" in name or "heat_rate" in name:
        return 100.0, 10.0, 15.0, 5.0, 0.0
    if "humidity" in name:
        return 40.0, 10.0, 0.0, 5.0, 0.0

    return 10.0, 2.0, 1.0, 1.0, 0.0


def _generate_chunk(cfg: GeneratorConfig, unit_ids: np.ndarray, rng: np.random.Generator) -> pd.DataFrame:
    n_units = unit_ids.size
    cycles = np.arange(1, cfg.cycles + 1, dtype=int)

    unit_id = np.repeat(unit_ids, cfg.cycles)
    cycle = np.tile(cycles, n_units)

    fail_min = max(int(cfg.cycles * 0.6), 5)
    failure_cycle = rng.integers(fail_min, cfg.cycles + 1, size=n_units)
    failure_cycle_full = np.repeat(failure_cycle, cfg.cycles)

    rul = (failure_cycle_full - cycle).clip(min=0).astype(int)
    fail_within_h = (rul <= cfg.horizon).astype(int)

    wear = cycle / failure_cycle_full
    unit_bias = rng.normal(0.0, 1.0, size=n_units)
    unit_bias_full = np.repeat(unit_bias, cfg.cycles)

    data: dict[str, np.ndarray] = {
        ID_COL: unit_id.astype(int),
        TIME_COL: cycle.astype(int),
        LABEL_FAILURE_CYCLE: failure_cycle_full.astype(int),
        LABEL_RUL: rul.astype(int),
        LABEL_FAIL_WITHIN_H: fail_within_h.astype(int),
    }

    for name in CANONICAL_FEATURES:
        if name == "operating_mode":
            mode = rng.integers(1, 4, size=n_units)
            data[name] = np.repeat(mode, cfg.cycles).astype(int)
            continue
        if name == "runtime_hours":
            data[name] = cycle.astype(float)
            continue

        mean, std, trend, noise, clip_low = _feature_spec(name)
        base = rng.normal(mean, std, size=n_units)
        base_full = np.repeat(base, cfg.cycles)
        values = base_full + trend * wear + (0.5 * unit_bias_full) + rng.normal(
            0.0, noise, size=unit_id.size
        )

        if name == "load_percent":
            values = np.clip(values, 0.0, 100.0)
        elif clip_low is not None:
            values = np.clip(values, clip_low, None)

        data[name] = values.astype(float)

    return pd.DataFrame(data)


def _write_outputs(
    cfg: GeneratorConfig,
    rng: np.random.Generator,
    out_dir: Path,
    base_name: str,
    fmt: str,
    excel_max_rows: int,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    csv_path = out_dir / f"{base_name}.csv"
    xlsx_path = out_dir / f"{base_name}.xlsx"

    if fmt in {"csv", "both"} and csv_path.exists():
        csv_path.unlink()

    use_excel = fmt in {"xlsx", "both"}
    use_csv = fmt in {"csv", "both"}

    if use_excel:
        try:
            from openpyxl import Workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required for Excel output. Install it first.") from exc
        if cfg.n_units * cfg.cycles > excel_max_rows:
            print(
                f"Excel row limit is {excel_max_rows:,}. Output will be split into multiple files.",
                file=sys.stderr,
            )
        rows_limit = excel_max_rows - 1  # reserve one row for header
        part = 1
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("data")
        rows_written = 0
    else:
        wb = None
        ws = None
        rows_limit = 0
        part = 1
        rows_written = 0

    def _flush_xlsx() -> Path:
        target = xlsx_path.with_name(f"{xlsx_path.stem}_part{part}{xlsx_path.suffix}")
        if wb is not None:
            wb.save(target)
        return target

    unit_ids = np.arange(1, cfg.n_units + 1, dtype=int)
    header_written = False

    for start in range(0, cfg.n_units, cfg.chunk_units):
        end = min(start + cfg.chunk_units, cfg.n_units)
        chunk_units = unit_ids[start:end]
        chunk = _generate_chunk(cfg, chunk_units, rng)

        if use_csv:
            chunk.to_csv(csv_path, mode="a", header=not header_written, index=False)

        if use_excel:
            if not header_written:
                ws.append(list(chunk.columns))
            for row in chunk.itertuples(index=False, name=None):
                if rows_written >= rows_limit:
                    written = _flush_xlsx()
                    print(f"Wrote Excel: {written}")
                    part += 1
                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet("data")
                    ws.append(list(chunk.columns))
                    rows_written = 0
                ws.append(row)
                rows_written += 1

        header_written = True

    if use_csv:
        print(f"Wrote CSV: {csv_path}")

    if use_excel and wb is not None:
        written = _flush_xlsx()
        print(f"Wrote Excel: {written}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a large synthetic predictive maintenance dataset.")
    parser.add_argument("--out-dir", default="data/generated", help="Output directory.")
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="csv")
    parser.add_argument("--n-units", type=int, default=2000)
    parser.add_argument("--cycles", type=int, default=1000)
    parser.add_argument("--horizon", type=int, default=30)
    parser.add_argument("--chunk-units", type=int, default=50)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    cfg = GeneratorConfig(
        n_units=args.n_units,
        cycles=args.cycles,
        horizon=args.horizon,
        chunk_units=args.chunk_units,
        seed=args.seed,
    )

    rng = np.random.default_rng(cfg.seed)

    out_dir = Path(args.out_dir)
    base_name = f"synthetic_pm_{cfg.n_units}x{cfg.cycles}"
    _write_outputs(cfg, rng, out_dir, base_name, args.format, excel_max_rows=EXCEL_MAX_ROWS)


if __name__ == "__main__":
    main()
